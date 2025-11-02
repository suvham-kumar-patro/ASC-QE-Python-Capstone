from openpyxl import load_workbook

def read_sheet(path, sheet_name=None):
    wb = load_workbook(path)
    if sheet_name is None:
        sheet = wb.active
    else:
        sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0] if rows else []
    data = []
    for r in rows[1:]:
        rowd = {headers[i]: r[i] for i in range(len(headers))}
        data.append(rowd)
    return data
